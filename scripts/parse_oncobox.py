#!/usr/bin/env python3
"""Parse OncoBox Brust N1.1.1 Excel specification.

Reads `input/data/oncobox-brust/OncoBoxBrust_N1.1.1_Spec.xlsx` and dumps
key sheets to stdout in a form useful for manual inspection and for
driving Logical Model / StructureMap creation.
"""
import openpyxl
import sys
import os

PATH = os.path.join(os.path.dirname(__file__), '..',
                    'input', 'data', 'oncobox-brust',
                    'OncoBoxBrust_N1.1.1_Spec.xlsx')


def dump_sheet(ws, max_rows=None, max_cols=None):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows and i >= max_rows:
            break
        rows.append(row[:max_cols] if max_cols else row)
    return rows


def main():
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
    mode = sys.argv[1] if len(sys.argv) > 1 else 'list'

    if mode == 'list':
        for s in wb.sheetnames:
            ws = wb[s]
            print(f'{s}: {ws.max_row} rows x {ws.max_column} cols')
        return

    if mode == 'header':
        for s in wb.sheetnames:
            ws = wb[s]
            print(f'=== {s} ===')
            for row in ws.iter_rows(max_row=3, values_only=True):
                print(row)
            print()
        return

    if mode == 'dump':
        sheet = sys.argv[2]
        max_rows = int(sys.argv[3]) if len(sys.argv) > 3 else None
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if max_rows and i >= max_rows:
                break
            print(f'{i}\t' + '\t'.join('' if v is None else str(v).replace('\n', ' | ').replace('\t', ' ')
                                      for v in row))
        return

    if mode == 'all':
        for s in wb.sheetnames:
            ws = wb[s]
            print(f'=========== SHEET: {s} ({ws.max_row} x {ws.max_column}) ===========')
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                print(f'{i}\t' + '\t'.join('' if v is None else str(v).replace('\n', ' | ').replace('\t', ' ')
                                          for v in row))
            print()
        return


if __name__ == '__main__':
    main()
